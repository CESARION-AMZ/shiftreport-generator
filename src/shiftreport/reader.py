"""Lectura de datos de entrada (CSV o Excel) a una lista de filas (dict)."""

from __future__ import annotations

import csv
from pathlib import Path


def _to_number(value: object) -> float | None:
    """Devuelve el valor como float si es numerico, o None si no lo es."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace(" ", "")
    try:
        return float(cleaned)
    except ValueError:
        cleaned = cleaned.replace(".", "").replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None


def resolve_input(path: str | Path) -> Path:
    """Si 'path' es una carpeta, devuelve el CSV/Excel MAS RECIENTE que contenga.

    Permite apuntar el tool a la carpeta donde tu sistema deja los exports y que
    coja automaticamente el ultimo, sin elegir archivo a mano.
    """
    p = Path(path)
    if p.is_dir():
        candidates = [
            f for f in p.iterdir()
            if f.is_file() and f.suffix.lower() in (".csv", ".xlsx", ".xlsm")
        ]
        if not candidates:
            raise FileNotFoundError(f"No hay archivos .csv/.xlsx en la carpeta: {p}")
        return max(candidates, key=lambda f: f.stat().st_mtime)
    return p


def load_rows(path: str | Path) -> tuple[list[str], list[dict]]:
    """Carga un archivo CSV o XLSX. Returns (headers, rows)."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {p}")
    suffix = p.suffix.lower()
    if suffix == ".csv":
        return _load_csv(p)
    if suffix in (".xlsx", ".xlsm"):
        return _load_excel(p)
    raise ValueError(f"Formato no soportado: '{suffix}'. Usa .csv o .xlsx.")


def _load_csv(p: Path) -> tuple[list[str], list[dict]]:
    with p.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        headers = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader]
    return headers, rows


def _load_excel(p: Path) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
    rows: list[dict] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        rows.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    wb.close()
    return headers, rows


def detect_numeric_columns(headers: list[str], rows: list[dict]) -> list[str]:
    """Detecta columnas cuyo contenido (no vacio) es siempre numerico."""
    numeric: list[str] = []
    for col in headers:
        values = [r.get(col) for r in rows]
        non_empty = [v for v in values if v not in (None, "")]
        if not non_empty:
            continue
        if all(_to_number(v) is not None for v in non_empty):
            numeric.append(col)
    return numeric
